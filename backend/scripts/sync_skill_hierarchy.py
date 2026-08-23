"""
Use Skill-Heirarchies-G6-G9.xlsx as a *naming reference*, then assign
Assessment-style Topic IDs to every curriculum chapter.

Pattern (from Excel examples):
  G{grade}_S{section}_{DOMAIN}_{FOCUS}
  e.g. G7_S1_PLA_DIVER, G6_S8_ELE_CIRCUITS

Excel rows are used when they clearly match a chapter.
All other chapters get the same style of ID (not g*_science_ch* and not G*_CH*).
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
XLSX = REPO / "Skill-Heirarchies-G6-G9.xlsx"
OUT_SKILLS = Path(__file__).resolve().parents[1] / "app" / "data" / "skill_hierarchy.json"
CURRICULUM = Path(__file__).resolve().parents[1] / "app" / "data" / "curriculum.json"


def _section_id(topic_id: str) -> str:
    m = re.match(r"^(G\d+_S\d+)_", topic_id)
    return m.group(1) if m else topic_id


def parse_excel(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grade = None
    core = None
    skills: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        core_cell, topic_id, curriculum_ref = (row + (None, None, None))[:3]
        if core_cell and str(core_cell).strip().upper().startswith("GRADE"):
            m = re.search(r"(\d+)", str(core_cell))
            grade = int(m.group(1)) if m else grade
            core = None
            continue
        if core_cell and str(core_cell).strip():
            core = str(core_cell).strip()
        tid = (str(topic_id).strip() if topic_id else "") or ""
        if not tid:
            continue
        skills.append(
            {
                "topic_id": tid,
                "section_id": _section_id(tid),
                "grade": grade,
                "core_concept": core,
                "curriculum_reference": (str(curriculum_ref).strip() if curriculum_ref else ""),
            }
        )
    return {
        "source": path.name,
        "description": (
            "Excel is the naming reference (G#_S#_DOMAIN_FOCUS). "
            "Curriculum topic_ids follow that pattern for every chapter so "
            "Chroma / Assessment / Analytics stay consistent."
        ),
        "skills": skills,
    }


# Full chapter Topic IDs — Excel examples preferred; rest invented in the same style.
# One primary ID per textbook chapter (S# = chapter number in that grade path).
LESSON_TOPIC_IDS: dict[str, tuple[str, ...]] = {
    # —— Grade 6 ——
    "g6_sci_01": ("G6_S1_ORG_CHARS", "G6_S1_ORG_CLASS"),  # Excel
    "g6_sci_02": ("G6_S2_MAT_STATES", "G6_S2_MAT_PROPS"),  # Excel
    "g6_sci_03": ("G6_S3_WAT_RESOUR", "G6_S3_WAT_CYCLE"),
    "g6_sci_04": ("G6_S4_ENE_SOURCES",),  # Excel
    "g6_sci_05": ("G6_S5_LIG_VISION", "G6_S5_LIG_EYE"),
    "g6_sci_06": ("G6_S6_SOU_HEARING", "G6_S6_SOU_WAVES"),
    "g6_sci_07": ("G6_S7_MAG_POLES", "G6_S7_MAG_FORCE"),
    "g6_sci_08": ("G6_S8_ELE_CIRCUITS", "G6_S8_ELE_CONDINS"),  # Excel
    "g6_sci_09": ("G6_S9_HEA_EFFECTS", "G6_S9_HEA_EXPAND"),
    "g6_sci_10": ("G6_S10_FOO_INTERAC", "G6_S10_FOO_CHAINS"),
    "g6_sci_11": ("G6_S11_WEA_CLIMATE", "G6_S11_WEA_PATTER"),
    # —— Grade 7 Part I (Excel examples) ——
    "g7_sci_01": ("G7_S1_PLA_DIVER", "G7_S1_PLA_CLASSIF"),
    "g7_sci_02": ("G7_S2_STA_CHARGES", "G7_S2_STA_CAPACIT"),
    "g7_sci_03": ("G7_S3_ELE_SOURCES", "G7_S3_ELE_CURRENTS"),
    "g7_sci_04": ("G7_S4_WAT_SOLVENT", "G7_S4_WAT_COOLANT"),
    "g7_sci_05": ("G7_S5_ACI_IDENTIF", "G7_S5_ACI_INDICAT"),
    "g7_sci_06": ("G7_S6_ANI_CLASSIF", "G7_S6_ANI_ADAPTAT"),
    "g7_sci_07": ("G7_S7_ENE_FORMS", "G7_S7_ENE_TRANSF"),
    "g7_sci_08": ("G7_S8_EAR_STRUCT", "G7_S8_EAR_TECTON"),
    "g7_sci_09": ("G7_S9_LIG_SHADOWS", "G7_S9_LIG_MIRRORS"),
    "g7_sci_10": ("G7_S10_MIC_LIGHT", "G7_S10_MIC_ELECTR"),
    # —— Grade 7 Part II (same style) ——
    "g7_sci_11": ("G7_S11_SOU_PROPAG", "G7_S11_SOU_HEARING"),
    "g7_sci_12": ("G7_S12_BIO_PROCESS", "G7_S12_BIO_CELLS"),
    "g7_sci_13": ("G7_S13_ATM_LAYERS", "G7_S13_ATM_WEATHER"),
    "g7_sci_14": ("G7_S14_HEA_TEMPER", "G7_S14_HEA_TRANSF"),
    "g7_sci_15": ("G7_S15_SOI_TYPES", "G7_S15_SOI_PROPERT"),
    "g7_sci_16": ("G7_S16_FOR_MOTION", "G7_S16_FOR_TYPES"),
    "g7_sci_17": ("G7_S17_FOO_NUTRIEN", "G7_S17_FOO_BALANC"),
    "g7_sci_18": ("G7_S18_MIN_ROCKS", "G7_S18_MIN_TYPES"),
    "g7_sci_19": ("G7_S19_ENE_SOURCES", "G7_S19_ENE_RENEW"),
    # —— Grade 8 Part I ——
    "g8_sci_01": ("G8_S1_BIO_DIVER", "G8_S1_MIC_IMPORT"),  # Excel BIO_DIVER as example
    "g8_sci_02": ("G8_S2_ANI_CLASSIF", "G8_S2_ANI_KEYS"),
    "g8_sci_03": ("G8_S3_PLA_PARTS", "G8_S3_PLA_FUNCT"),
    "g8_sci_04": ("G8_S4_MAT_ELEMENTS", "G8_S4_MAT_COMPOUNDS"),  # Excel
    "g8_sci_05": ("G8_S5_SOU_WAVES", "G8_S5_SOU_PROPAG"),
    "g8_sci_06": ("G8_S6_MAG_FORCE", "G8_S6_MAG_FIELD"),
    "g8_sci_07": ("G8_S7_ELE_MEASURE", "G8_S7_ELE_UNITS"),
    "g8_sci_08": ("G8_S8_CHA_PHYSICAL", "G8_S8_CHA_BURNING"),  # Excel CHA_* style
    # —— Grade 8 Part II ——
    "g8_sci_09": ("G8_S9_SYS_HUMAN", "G8_S9_SYS_ORGANS"),
    "g8_sci_10": ("G8_S10_ELE_CIRCUIT", "G8_S10_ELE_SAFETY"),
    "g8_sci_11": ("G8_S11_PHO_PROCESS", "G8_S11_PHO_IMPORT"),  # Excel PHO_* style
    "g8_sci_12": ("G8_S12_LIF_CYCLES", "G8_S12_LIF_STAGES"),
    "g8_sci_13": ("G8_S13_FOO_PRESERV", "G8_S13_FOO_METHODS"),
    "g8_sci_14": ("G8_S14_SOL_SYSTEM", "G8_S14_SOL_EXPLORE"),
    "g8_sci_15": ("G8_S15_DIS_NATURAL", "G8_S15_DIS_SAFETY"),
    # —— Grade 9 Part I ——
    "g9_sci_01": ("G9_S1_MIC_APPLIC", "G9_S1_MIC_USES"),
    "g9_sci_02": ("G9_S2_SEN_EYE", "G9_S2_SEN_EAR"),
    "g9_sci_03": ("G9_S3_NAT_ATOMS", "G9_S3_NAT_CONFIG"),  # Excel NAT_* style
    "g9_sci_04": ("G9_S4_FOR_BASIC", "G9_S4_FOR_LAWS"),
    "g9_sci_05": ("G9_S5_PRE_SOLID", "G9_S5_PRE_CALC"),
    "g9_sci_06": ("G9_S6_SYS_CIRCUL", "G9_S6_SYS_BLOOD"),  # Excel SYS_CIRCUL style
    "g9_sci_07": ("G9_S7_PLA_GROWTH", "G9_S7_PLA_HORMON"),
    "g9_sci_08": ("G9_S8_ORG_SUPPORT", "G9_S8_ORG_MOVE"),
    "g9_sci_09": ("G9_S9_EVO_PROCESS", "G9_S9_EVO_EVIDEN"),
    # —— Grade 9 Part II ——
    "g9_sci_10": ("G9_S10_ELE_LYSIS", "G9_S10_ELE_IONS"),
    "g9_sci_11": ("G9_S11_MAT_DENSITY", "G9_S11_MAT_MEASURE"),  # Excel MAT_DENSITY style
    "g9_sci_12": ("G9_S12_BIO_DIVER", "G9_S12_BIO_CONSERV"),
    "g9_sci_13": ("G9_S13_ENV_GREEN", "G9_S13_ENV_ARTIFIC"),
    "g9_sci_14": ("G9_S14_LIG_REFRAC", "G9_S14_LIG_REFLECT"),  # Excel LIG_REFRAC style
    "g9_sci_15": ("G9_S15_MAC_SIMPLE", "G9_S15_MAC_WORK"),
    "g9_sci_16": ("G9_S16_NAN_TECH", "G9_S16_NAN_APPLIC"),
    "g9_sci_17": ("G9_S17_LIG_ACCIDEN", "G9_S17_LIG_SAFETY"),  # lightning
    "g9_sci_18": ("G9_S18_DIS_NATURAL", "G9_S18_DIS_PREPARE"),
    "g9_sci_19": ("G9_S19_NAT_SUSTAIN", "G9_S19_NAT_RESOUR"),
}


def _legacy_topic(lesson_id: str) -> str:
    m = re.search(r"^g(\d+)_sci_(\d+)$", lesson_id)
    if not m:
        return ""
    return f"g{m.group(1)}_science_ch{int(m.group(2)):02d}"


def update_curriculum(skills_doc: dict) -> dict:
    raw = json.loads(CURRICULUM.read_text(encoding="utf-8"))
    mapping: dict[str, dict] = {}
    skill_to_lesson: dict[str, str] = {}

    for book in raw.get("books") or []:
        grade = book.get("grade")
        for le in book.get("lessons") or []:
            lid = le["lesson_id"]
            title = le.get("title") or lid
            if lid not in LESSON_TOPIC_IDS:
                raise SystemExit(f"Missing Topic ID mapping for {lid} ({title})")
            skill_ids = list(LESSON_TOPIC_IDS[lid])
            primary = skill_ids[0]
            legacy = le.get("legacy_topic_id") or _legacy_topic(lid)
            # Don't treat already-new IDs as legacy
            if legacy.startswith("G") and "_S" in legacy:
                legacy = _legacy_topic(lid)

            le["topic_id"] = primary
            le["legacy_topic_id"] = legacy
            le["skill_topic_ids"] = skill_ids
            le["skill_section_id"] = _section_id(primary)

            mapping[lid] = {
                "lesson_id": lid,
                "title": title,
                "topic_id": primary,
                "legacy_topic_id": legacy,
                "skill_topic_ids": skill_ids,
                "skill_section_id": le["skill_section_id"],
                "grade": grade,
                "from_excel_example": any(
                    s["topic_id"] == primary for s in skills_doc.get("skills") or []
                ),
            }
            for sid in skill_ids:
                skill_to_lesson.setdefault(sid, lid)
            skill_to_lesson.setdefault(primary, lid)
            if legacy:
                skill_to_lesson.setdefault(legacy, lid)

    skills_doc["lesson_mappings"] = mapping
    skills_doc["skill_to_lesson"] = skill_to_lesson
    skills_doc["naming_pattern"] = "G{grade}_S{section}_{DOMAIN}_{FOCUS}"
    CURRICULUM.write_text(json.dumps(raw, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return skills_doc


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Missing Excel: {XLSX}")
    doc = parse_excel(XLSX)
    doc = update_curriculum(doc)
    OUT_SKILLS.parent.mkdir(parents=True, exist_ok=True)
    OUT_SKILLS.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_SKILLS}")
    print(f"Updated {CURRICULUM}")
    excel_hits = sum(1 for m in doc["lesson_mappings"].values() if m.get("from_excel_example"))
    print(f"Lessons: {len(doc['lesson_mappings'])} (Excel example used as primary: {excel_hits})")
    for lid in ("g7_sci_01", "g6_sci_03", "g7_sci_11", "g8_sci_05", "g9_sci_15"):
        m = doc["lesson_mappings"][lid]
        print(f"  {lid}: {m['legacy_topic_id']} -> {m['topic_id']}  skills={m['skill_topic_ids']}")


if __name__ == "__main__":
    main()
